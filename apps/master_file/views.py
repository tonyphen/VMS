from django.contrib.auth.decorators import user_passes_test
from django.conf import settings
from django.db.models import Count
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import load_workbook
from django.utils.html import escape

from apps.basic.utilities import created_updated
from apps.basic import dictionary
from apps.master_file import models, forms


def wood_definition(request):
    qsProfile_Master = models.ProfileMaster.objects.all()
    qsProfile = models.Profile.objects.all()
    qsWoodType = models.WoodType.objects.all().order_by('wood_type_id')
    qsSortGroup = models.SortGroup.objects.all().order_by('id')
    qsLength = models.Length.objects.all().order_by('feet')

    pageLength = settings.PAGE_LENGTH
    context = {
        'profile_master_list': qsProfile_Master,
        'profile_list': qsProfile,
        'wood_type_list': qsWoodType,
        'soft_group_list': qsSortGroup,
        'length_list': qsLength,
        'pageLength': pageLength
    }
    return render(request, 'master_file/wood_definition.html', context)


def master_file_upload(request):
    if request.method == 'POST':
        file_name = request.FILES['fileName'].name
        upload_file = request.FILES['fileName']
        workbook = load_workbook(upload_file, read_only=True)
        # Get name of the first sheet and open the sheet name
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)
        data = []
        if str(file_name).lower() == 'profile.master.xlsx':
            for row in worksheet.iter_rows(min_row=2):
                pmaster = models.ProfileMaster()
                pmaster.group_id = row[0].value
                pmaster.group_name = row[1].value
                pmaster.sales_target = int(row[2].value)
                pmaster.created_by = request.user
                pmaster.updated_by = request.user
                data.append(pmaster)
            models.ProfileMaster.objects.bulk_create(data)
            return redirect('wood_definition')

        if str(file_name).lower() == 'length_master.xlsx':
            for row in worksheet.iter_rows(min_row=2):
                length = models.Length()
                length.feet = round(row[0].value, 4)
                length.inch = round(row[1].value, 2)
                length.mm = int(row[2].value)
                if row[3].value == 1:
                    length.freqUsed = True
                else:
                    length.freqUsed = False
                length.created_by = request.user
                length.updated_by = request.user
                data.append(length)
            models.Length.objects.bulk_create(data)
            return redirect('wood_definition')

        if str(file_name).lower() == 'profiles.xlsx':
            for row in worksheet.iter_rows(min_row=2):
                profile = models.Profile()
                profile.group_master = models.ProfileMaster.objects.get(group_id=row[0].value)
                profile.profile_group = row[1].value
                profile.profile_group_description = str(row[2].value).upper()
                profile.profile_id = str(row[3].value).upper()
                profile.profile_sym = str(row[4].value).upper()
                profile.description = str(row[5].value).upper()
                profile.image = None
                profile.width = round(float(row[6].value), 2)
                profile.thick = round(float(row[7].value), 3)
                profile.CBM_Feet = round(float(row[8].value), 6)
                profile.M2_Feet = round(float(row[9].value), 6)
                profile.bundle = int(row[10].value)
                profile.created_by = request.user
                profile.updated_by = request.user
                data.append(profile)
            models.Profile.objects.bulk_create(data)
            return redirect('wood_definition')

        if str(file_name).lower() == 'sortgroup.xlsx':
            for row in worksheet.iter_rows(min_row=2):
                sortgroup = models.SortGroup()
                sortgroup.id = row[0].value
                sortgroup.description = row[1].value
                sortgroup.sym = str(row[2].value).upper()
                sortgroup.parent = row[3].value
                if row[4].value == 1:
                    freqused = True
                else:
                    freqused = False

                sortgroup.freqUse = freqused
                sortgroup.created_by = request.user
                sortgroup.updated_by = request.user
                data.append(sortgroup)
            models.SortGroup.objects.bulk_create(data)
            return redirect('wood_definition')

        if str(file_name).lower() == 'woodtype.xlsx':
            for row in worksheet.iter_rows(min_row=2):
                wt = models.WoodType()
                wt.wood_type_id = row[0].value
                wt.description = str(row[1].value).upper()
                wt.parent_id = row[2].value
                if row[3].value == 1:
                    freqused = True
                else:
                    freqused = False
                wt.freqUsed = freqused
                wt.wood_group = row[4].value
                wt.prod_type = row[5].value
                wt.sym = str(row[6].value).upper()
                if row[7].value == 1:
                    bodview = True
                else:
                    bodview = False
                wt.bod_view = bodview
                wt.created_by = request.user
                wt.updated_by = request.user
                data.append(wt)
            models.WoodType.objects.bulk_create(data)
            return redirect('wood_definition')

        if str(file_name).lower() == 'thick.xlsx':
            for row in worksheet.iter_rows(min_row=2):
                thick = models.Thick()
                thick.mm = row[0].value
                thick.cm = row[1].value
                thick.inch = row[2].value
                thick.created_by = request.user
                thick.updated_by = request.user
                data.append(thick)
            models.Thick.objects.bulk_create(data)
            return redirect('size')

        if str(file_name).lower() == 'log_scale.xlsx':
            for row in worksheet.iter_rows(min_row=2):
                log = models.LogScale()
                log.dia = row[0].value
                log.length = row[1].value
                log.BF = row[2].value
                data.append(log)
            models.LogScale.objects.bulk_create(data)
            return redirect('size')

        if str(file_name).lower() == 'fg_products.xlsx':
            for row in worksheet.iter_rows(min_row=2):
                fg = models.Product()
                fg.warehouse_id = row[0].value
                fg.main_category_id = row[1].value
                fg.parent_category = row[2].value
                fg.code = row[4].value
                fg.description = row[5].value
                fg.profile_id = row[6].value[1:5]
                fg.wood_type_id = row[7].value[1:3]
                fg.unit_id = row[8].value
                fg.min_qty = 0
                fg.max_qty = 0
                fg.created_by_id = row[11].value
                fg.updated_by_id = row[12].value

                fg.save()

            return redirect('product_list', main_cat='FINISHED_GOODS')

        if str(file_name).lower() == 'colors.xlsx':
            for row in worksheet.iter_rows(min_row=2):
                color = models.Color()
                color.color_group = row[0].value
                # if row[5].value:
                #     print("'" + row[5].value[1:3] + "'")
                color.color_id = row[1].value[1:5]
                color.description = row[2].value
                color.sort_group_id = row[3].value
                color.sort_group_note = row[4].value
                if row[5].value:
                    color.wood_type_id = row[5].value[1:3]
                else:
                    color.wood_type_id = None
                color.gloss = row[6].value
                color.printed = row[7].value
                color.distressed = row[8].value
                color.distressed_remark = row[9].value
                color.phun_hot = row[10].value
                color.emboss = row[11].value
                color.scratch = row[12].value
                color.glazed = row[13].value
                color.danh_bui = row[14].value
                color.remark = row[15].value
                color.created_by_id = 1
                color.updated_by_id = 1
                data.append(color)
                # color.save()
            models.Color.objects.bulk_create(data)
            return redirect('color_list')

    return render(request, 'master_file/master_file_upload.html')


def size(request):
    qsLength = models.Length.objects.all()
    qsThick = models.Thick.objects.all()
    qsLogScale = models.LogScale.objects.all()

    context = {
        'length_list': qsLength,
        'thick_list': qsThick,
        'log_scale_list': qsLogScale,
        'pageLength': settings.PAGE_LENGTH
    }
    return render(request, 'master_file/size.html', context)


def category_list(request):
    qsCat = models.Category.objects.filter(is_root=True)
    context = {
        'cat_list': qsCat
    }
    return render(request, 'master_file/category_list.html', context)


def profile_master_create(request):
    form = forms.ProfileMasterForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.instance.created_by = request.user
            form.instance.updated_by = request.user
            form.save()
            return redirect('wood_definition')
    context = {
        'form_name': 'Profile Master',
        'parent_link': "{% url 'wood_definition' %}",
        'parent_name': 'Wood definitions',
        'form': form,
    }
    return render(request, 'layouts/single_form.html', context)


def profile_master_update(request, id):
    obj = get_object_or_404(models.ProfileMaster, group_id=id)
    form = forms.ProfileMasterForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.instance.updated_by = request.user
        form.save()
        return redirect('wood_definition')
    context = {
        'form_name': 'Profile Master',
        'parent_link': "{% url 'wood_definition' %}",
        'parent_name': 'Wood definitions',
        'form': form,
    }
    return render(request, 'layouts/single_form.html', context)


def profile_master_delete(request, id):
    obj = get_object_or_404(models.ProfileMaster, group_id=id)
    if request.method == "POST":
        obj.delete()
        return redirect('wood_definition')
    return render(request, 'layouts/page-404.html')


def profile_create(request):
    form = forms.ProfileForm(request.POST or None)
    if request.method == 'POST':
        form.instance.created_by = request.user
        if form.is_valid():
            form.save()
            created_updated(models.Profile, request)
            return redirect('wood_definition')
    context = {
        'form_name': 'Profile',
        'parent_link': "{% url 'wood_definition' %}",
        'parent_name': 'Wood definitions',
        'form': form,
    }
    return render(request, 'layouts/single_form.html', context)


def profile_update(request, id):
    obj = get_object_or_404(models.Profile, profile_id=id)
    form = forms.ProfileForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.instance.updated_by = request.user
        form.save()
        return redirect('wood_definition')
    context = {
        'form_name': 'Profile',
        'parent_link': "{% url 'wood_definition' %}",
        'parent_name': 'Wood definitions',
        'form': form,
    }
    return render(request, 'layouts/single_form.html', context)


def profile_delete(request, id):
    obj = get_object_or_404(models.Profile, profile_id=id)
    if request.method == "POST":
        obj.delete()
        return redirect('wood_definition')
    return render(request, 'layouts/page-404.html')


def wood_type_create(request):
    form = forms.WoodTypeForm(request.POST or None)
    if request.method == 'POST':
        form.instance.created_by = request.user
        if form.is_valid():
            form.save()
            created_updated(models.WoodType, request)
            return redirect('wood_definition')
    context = {
        'form_name': 'Wood Type',
        'parent_link': "{% url 'wood_definition' %}",
        'parent_name': 'Wood definitions',
        'form': form,
    }
    return render(request, 'layouts/single_form.html', context)


def wood_type_update(request, id):
    obj = get_object_or_404(models.WoodType, wood_type_id=id)
    form = forms.WoodTypeForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.instance.updated_by = request.user
        form.save()
        return redirect('wood_definition')
    context = {
        'form_name': 'Wood Type',
        'parent_link': "{% url 'wood_definition' %}",
        'parent_name': 'Wood definitions',
        'form': form,
    }
    return render(request, 'layouts/single_form.html', context)


def wood_type_delete(request, id):
    obj = get_object_or_404(models.WoodType, wood_type_id=id)
    if request.method == "POST":
        obj.delete()
        return redirect('wood_definition')
    return render(request, 'layouts/page-404.html')


def sort_group_create(request):
    form = forms.SortGroupForm(request.POST or None)
    if request.method == 'POST':
        form.instance.created_by = request.user
        if form.is_valid():
            form.save()
            created_updated(models.SortGroup, request)
            return redirect('wood_definition')
    context = {
        'form_name': 'Sort Group',
        'parent_link': "{% url 'wood_definition' %}",
        'parent_name': 'Wood definitions',
        'form': form,
    }
    return render(request, 'layouts/single_form.html', context)


def sort_group_update(request, id):
    obj = get_object_or_404(models.SortGroup, id=id)
    form = forms.SortGroupForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.instance.updated_by = request.user
        form.save()
        return redirect('wood_definition')
    context = {
        'form_name': 'Sort Group',
        'parent_link': "{% url 'wood_definition' %}",
        'parent_name': 'Wood definitions',
        'form': form,
    }
    return render(request, 'layouts/single_form.html', context)


def sort_group_delete(request, id):
    obj = get_object_or_404(models.SortGroup, id=id)
    if request.method == "POST":
        obj.delete()
        return redirect('wood_definition')
    return render(request, 'layouts/page-404.html')


def length_create(request):
    form = forms.LengthForm(request.POST or None)
    if request.method == 'POST':
        form.instance.created_by = request.user
        if form.is_valid():
            form.instance.created_by = request.user
            form.instance.updated_by = request.user
            form.save()
            return redirect('wood_definition')
    context = {
        'form_name': 'Length',
        'parent_link': "{% url 'wood_definition' %}",
        'parent_name': 'Wood definitions',
        'form': form,
    }
    return render(request, 'layouts/single_form.html', context)


def length_update(request, id):
    obj = get_object_or_404(models.Length, mm=id)
    form = forms.LengthForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.instance.updated_by = request.user
        form.save()
        return redirect('wood_definition')
    context = {
        'form_name': 'Length',
        'parent_link': "{% url 'wood_definition' %}",
        'parent_name': 'Wood definitions',
        'form': form,
    }
    return render(request, 'layouts/single_form.html', context)


def length_delete(request, id):
    obj = get_object_or_404(models.Length, mm=id)
    if request.method == "POST":
        obj.delete()
        return redirect('wood_definition')
    return render(request, 'layouts/page-404.html')


def thick_create(request):
    form = forms.ThickForm(request.POST or None)
    if request.method == 'POST':
        form.instance.created_by = request.user
        if form.is_valid():
            form.instance.created_by = request.user
            form.instance.updated_by = request.user
            form.save()
            return redirect('size')
    context = {
        'form_name': 'Thick',
        'parent_link': "{% url 'size' %}",
        'parent_name': 'Size master file',
        'form': form,
    }
    return render(request, "layouts/single_form.html", context)


def thick_update(request, id):
    obj = get_object_or_404(models.Thick, mm=id)
    form = forms.ThickForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.instance.updated_by = request.user
        form.save()
        return redirect('size')
    context = {
        'form_name': 'Thick',
        'parent_link': "{% url 'size' %}",
        'parent_name': 'Size master file',
        'form': form,
    }
    return render(request, "layouts/single_form.html", context)


def thick_delete(request, id):
    obj = get_object_or_404(models.Thick, mm=id)
    if request.method == "POST":
        obj.delete()
        return redirect('size')
    return render(request, 'layouts/page-404.html')


def log_scale_create(request):
    form = forms.LogScaleForm(request.POST or None)
    if request.method == 'POST':
        form.instance.created_by = request.user
        if form.is_valid():
            form.instance.created_by = request.user
            form.instance.updated_by = request.user
            form.save()
            return redirect('size')
    context = {
        'form_name': 'Log Scale',
        'parent_link': "{% url 'size' %}",
        'parent_name': 'Size master file',
        'form': form,
    }
    return render(request, "layouts/single_form.html", context)


def log_scale_update(request, id):
    obj = get_object_or_404(models.LogScale, id=id)
    form = forms.LogScaleForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.instance.updated_by = request.user
        form.save()
        return redirect('size')
    context = {
        'form_name': 'Log Scale',
        'parent_link': "{% url 'size' %}",
        'parent_name': 'Size master file',
        'form': form,
    }
    return render(request, "layouts/single_form.html", context)


def log_scale_delete(request, id):
    obj = get_object_or_404(models.LogScale, id=id)
    if request.method == "POST":
        obj.delete()
        return redirect('size')
    return render(request, 'layouts/page-404.html')


def warehouse_list(request):
    qsWarehouse = models.Warehouse.objects.all()
    context = {
        'warehouse_list': qsWarehouse,
        'pageLength': settings.PAGE_LENGTH
    }
    return render(request, 'master_file/warehouse_list.html', context)


def warehouse_create(request):
    form = forms.WarehouseForm(request.POST or None, id_readonly=False)
    if request.method == 'POST':
        form.instance.created_by = request.user
        if form.is_valid():
            form.instance.id = str(form.instance.id).upper()
            form.instance.description = str(form.instance.description).upper()
            form.save()
            return redirect('warehouse_list')
    context = {
        'form_name': 'Warehouse',
        'form': form,
    }
    return render(request, "layouts/single_form.html", context)


def warehouse_update(request, id):
    obj = get_object_or_404(models.Warehouse, id=id)
    form = forms.WarehouseForm(request.POST or None, id_readonly=True, instance=obj)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('warehouse_list')
    else:
        context = {
            'form_name': 'Warehouse',
            'form': form,
        }
        return render(request, "layouts/single_form.html", context)


def warehouse_delete(request, id):
    obj = get_object_or_404(models.Warehouse, id=id)
    if request.method == "POST":
        obj.delete()
        return redirect('warehouse_list')
    return render(request, 'layouts/page-404.html')


def product_list(request, main_cat):
    context = {
        'main_cat': main_cat,
        'pageLength': settings.LIST_LENGTH
    }
    return render(request, 'master_file/product_list.html', context)


def product_create(request, main_cat):

    if main_cat == dictionary._FG_PRODUCT:
        frm = forms.FGProductForm
    elif main_cat == dictionary._RAW_PRODUCT:
        frm = forms.RWProductForm

    form = frm(request.POST or None, main_cat=main_cat)
    if request.method == 'POST':
        print('F' + request.POST.get('profile') + request.POST.get('wood_type'))
        if form.is_valid():
            # Tạo code FG
            # if main_cat == dictionary._FG_PRODUCT:

            form.instance.created_by = request.user
            form.instance.updated_by = request.user
            form.save()
            return redirect('product_list', main_cat=main_cat)
    context = {
        'form_name': 'Create a Product',
        'parent_link': '/mf/' + main_cat + '/product_list/',
        'parent_name': 'Product List',
        'main_cat': main_cat,
        'form': form,
    }
    return render(request, "layouts/single_form.html", context)


def product_update(request, main_cat, id):
    obj = get_object_or_404(models.Product, category=main_cat, code=id)
    form = forms.FGProductForm(request.POST or None, main_cat=main_cat, instance=obj)

    if form.is_valid():
        form.instance.updated_by = request.user
        form.save()
        return redirect('product_list', main_cat)
    context = {
        'form_name': 'Edit Product',
        'parent_link': '/mf/' + main_cat + '/product_list/',
        'parent_name': 'Product List',
        'main_cat': main_cat,
        'form': form,
    }
    return render(request, "layouts/single_form.html", context)


def product_delete(request, id):
    obj = get_object_or_404(models.Product, code=id)
    if request.method == "POST":
        obj.delete()
        return redirect('product_list')
    return render(request, 'layouts/page-404.html')


def color_list(request):
    context = {
        'pageLength': settings.LIST_LENGTH
    }
    return render(request, 'master_file/color_list.html', context)


def color_create(request):
    form = forms.ColorForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.instance.created_by = request.user
            form.instance.updated_by = request.user
            form.save()
            return redirect('color_list')
    context = {
        'form_name': 'Color List',
        # 'parent_link': "{% url '' %}",
        # 'parent_name': '',
        'form': form,
    }
    return render(request, "layouts/single_form.html", context)


def color_update(request, id):
    if not request.user.has_perm('can_edit_color'):
        return render(request, 'layouts/no_permission.html')
    obj = get_object_or_404(models.Color, color_id=id)
    form = forms.ColorForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.instance.updated_by = request.user
        form.save()
        return redirect('color_list')
    context = {
        'form_name': 'Color List',
        'parent_link': '/mf/color_list/',
        'parent_name': 'Color List',
        'form': form,
    }
    return render(request, "layouts/single_form.html", context)


def color_delete(request, id):
    obj = get_object_or_404(models.Color, color_id=id)
    if request.method == "POST":
        obj.delete()
        return redirect('color')
    return render(request, 'layouts/page-404.html')

